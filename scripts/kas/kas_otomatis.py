#!/usr/bin/env python3
"""
kas_otomatis.py
Laporan Kas Bulanan Otomatis - PT Adhitama Persada Indonesia
─────────────────────────────────────────────────────────────
Jalankan setiap tanggal 2 via GitHub Actions.
Script akan:
  1. Download XLSX terbaru dari Google Sheets (public export)
  2. Parsing semua tab SUMMARY untuk ekstrak data per bulan
  3. Hitung bulan lalu + YTD Tahun Fiskal (mulai April)
  4. Kirim ringkasan 3 bagian via WhatsApp (Fonnte API)
  5. Tulis log hasil eksekusi
"""

import os, sys, re, requests, tempfile, traceback, logging
from datetime import date, datetime, timezone, timedelta

# ─── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
LOG_PATH    = os.path.join(SCRIPT_DIR, 'kas_otomatis.log')

# ─── Config dari environment variables (bukan file lagi) ─────────────────────
FONNTE_TOKEN   = os.environ["FONNTE_TOKEN"]
WA_TARGET      = os.environ["KAS_WA_TARGET"]
SPREADSHEET_ID = os.environ["KAS_SPREADSHEET_ID"]
COMPANY_NAME   = os.environ.get("KAS_COMPANY_NAME", "PT Adhitama Persada Indonesia")
FY_START_MONTH = int(os.environ.get("KAS_FISCAL_YEAR_START_MONTH", "4"))   # April
DOWNLOAD_URL   = (
    f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}"
    f"/export?format=xlsx"
)

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s  %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(LOG_PATH, encoding='utf-8'),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger(__name__)

# ─── Indonesian month names ───────────────────────────────────────────────────
BULAN_ID = {
    'januari': 1,  'februari': 2,  'maret': 3,    'april': 4,
    'mei': 5,      'juni': 6,      'juli': 7,      'agustus': 8,
    'september': 9,'sepetember': 9,
    'oktober': 10, 'november': 11, 'desember': 12,
}
BULAN_NAMA = {
    1:'Januari', 2:'Februari', 3:'Maret',    4:'April',
    5:'Mei',     6:'Juni',     7:'Juli',      8:'Agustus',
    9:'September',10:'Oktober',11:'November',12:'Desember',
}

# ─── Helper: safe number ──────────────────────────────────────────────────────
def get_num(v):
    try:
        return float(str(v).replace(',', '').replace(' ', ''))
    except Exception:
        return 0.0

# ─── Format nilai dalam Jt ───────────────────────────────────────────────────
def fmt(jt):
    """'1.41M' jika >= 1000Jt, '590Jt' jika < 1000Jt"""
    if jt >= 1000:
        return f"{jt/1000:.2f}M"
    return f"{round(jt)}Jt"

# ─── Parse tab name → (month, year) ──────────────────────────────────────────
def parse_tab_month_year(tab_name: str):
    name   = tab_name.lower()
    year_m = re.search(r'(202\d)', name)
    year   = int(year_m.group(1)) if year_m else None
    month  = None
    for mn in sorted(BULAN_ID, key=len, reverse=True):
        if mn in name:
            month = BULAN_ID[mn]
            break
    if month and not year:
        year = 2024
    return (month, year) if (month and year) else None

# ─── Fiscal-year helpers ──────────────────────────────────────────────────────
def fiscal_year_of(d: date) -> int:
    return d.year if d.month >= FY_START_MONTH else d.year - 1

def prev_month(d: date) -> date:
    if d.month == 1:
        return d.replace(year=d.year - 1, month=12, day=1)
    return d.replace(month=d.month - 1, day=1)

def next_month(d: date) -> date:
    if d.month == 12:
        return d.replace(year=d.year + 1, month=1, day=1)
    return d.replace(month=d.month + 1, day=1)

# ─── Download XLSX ────────────────────────────────────────────────────────────
def download_spreadsheet() -> str:
    log.info("Downloading spreadsheet from Google Sheets...")
    resp = requests.get(DOWNLOAD_URL, timeout=90)
    resp.raise_for_status()
    tmp = tempfile.mktemp(suffix='.xlsx')
    with open(tmp, 'wb') as f:
        f.write(resp.content)
    log.info(f"Downloaded {len(resp.content)/1024:.1f} KB → {tmp}")
    return tmp

# ─── Extract monthly data ─────────────────────────────────────────────────────
def extract_monthly_data(xlsx_path: str) -> dict:
    import openpyxl
    log.info("Loading workbook...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    monthly = {}

    for tab_name in wb.sheetnames:
        if 'SUMMARY' not in tab_name.upper():
            continue
        my = parse_tab_month_year(tab_name)
        if not my:
            log.warning(f"  Skipping (can't parse date): {tab_name}")
            continue

        month, year = my
        ws = wb[tab_name]

        total_expense = 0.0
        saldo_masuk   = 0.0
        saldo         = 0.0

        for row in ws.iter_rows(values_only=True):
            r  = list(row) + [None] * 10
            v0 = str(r[0]).strip() if r[0] else ''
            v1 = str(r[1]).strip() if r[1] else ''
            v2 = str(r[2]).strip() if r[2] else ''

            if 'TOTAL' in v0.upper() and v2:
                total_expense = get_num(v2)
            if 'Saldo Masuk' in v1 and v2:
                saldo_masuk = get_num(v2)
            if v1 == 'Saldo' and v2:
                saldo = get_num(v2)
            if v1 == 'Pengeluaran ' and v2:
                total_expense = get_num(v2)

        key      = (year, month)
        existing = monthly.get(key)
        if existing and existing['income'] > 0 and saldo_masuk == 0:
            continue

        if saldo_masuk == 0:
            saldo_masuk = extract_kas_income(wb, month, year)

        monthly[key] = {
            'income':  saldo_masuk,
            'expense': total_expense,
            'saldo':   saldo,
            'tab':     tab_name,
        }
        log.info(
            f"  {tab_name:40s} → "
            f"income={saldo_masuk:>14,.0f}  "
            f"expense={total_expense:>14,.0f}"
        )

    wb.close()
    log.info(f"Total months parsed: {len(monthly)}")
    return monthly

def extract_kas_income(wb, month: int, year: int) -> float:
    target = (month, year)
    for tab_name in wb.sheetnames:
        if 'KAS' not in tab_name.upper() and 'HO' not in tab_name.upper():
            continue
        if 'SUMMARY' in tab_name.upper():
            continue
        if parse_tab_month_year(tab_name) != target:
            continue
        ws = wb[tab_name]
        total_kredit = 0.0
        for row in ws.iter_rows(values_only=True):
            r = list(row) + [None] * 10
            if r[4] not in ('Kredit +', 'Kredit', None, ''):
                v = get_num(r[4])
                if v > 0:
                    total_kredit += v
        if total_kredit > 0:
            return total_kredit
    return 0.0

# ─── Analisa otomatis ─────────────────────────────────────────────────────────
def build_analisa(ytd_data: list) -> str:
    """
    ytd_data: list of (bln_short, year, masuk_jt, keluar_jt).
    Hasilkan analisa singkat berdasarkan tren surplus/defisit.
    """
    if not ytd_data:
        return ""

    surpluses = [(b, y, m - k) for b, y, m, k in ytd_data]
    ytd_sur   = sum(s for _, _, s in surpluses)

    lines = []

    if len(surpluses) >= 2:
        prev_bln, prev_yr, prev_sur = surpluses[-2]
        last_bln, last_yr, last_sur = surpluses[-1]

        if last_sur < 0 and prev_sur < 0:
            if abs(last_sur) > abs(prev_sur):
                tren = f"⚠️ Defisit melebar ({fmt(abs(prev_sur))} → {fmt(abs(last_sur))})"
                rec  = "Segera tinjau pos pengeluaran terbesar dan upayakan percepatan penagihan."
            else:
                tren = f"📉 Defisit mengecil ({fmt(abs(prev_sur))} → {fmt(abs(last_sur))})"
                rec  = "Tren membaik — pertahankan kontrol pengeluaran."
        elif last_sur >= 0 and prev_sur >= 0:
            if last_sur < prev_sur:
                tren = f"⚠️ Surplus menurun ({fmt(prev_sur)} → {fmt(last_sur)})"
                rec  = "Perlu efisiensi pengeluaran atau tingkatkan pendapatan sebelum surplus habis."
            else:
                tren = f"✅ Surplus membaik ({fmt(prev_sur)} → {fmt(last_sur)})"
                rec  = "Pertahankan momentum — jaga kontrol pengeluaran."
        elif last_sur < 0 and prev_sur >= 0:
            tren = f"🔴 Berbalik defisit (+{fmt(prev_sur)} → ({fmt(abs(last_sur))}))"
            rec  = "Waspadai lonjakan pengeluaran — evaluasi anggaran segera."
        else:
            tren = f"🟢 Berbalik surplus (({fmt(abs(prev_sur))}) → +{fmt(last_sur)})"
            rec  = "Pemulihan positif — jaga agar tetap surplus di bulan berikutnya."
    else:
        tren = "Data baru 1 bulan di tahun fiskal ini"
        rec  = "Pantau terus di bulan berikutnya."

    ytd_ket = (f"✅ YTD masih surplus {fmt(ytd_sur)}"
               if ytd_sur >= 0
               else f"🔴 YTD defisit {fmt(abs(ytd_sur))}")

    lines.append(f"📌 *Tren:* {tren}")
    lines.append(f"📌 *YTD:* {ytd_ket}")
    lines.append(f"💡 _{rec}_")
    return "\n".join(lines)

# ─── Build WA message (3 bagian) ─────────────────────────────────────────────
def build_wa_message(monthly: dict, reference_date: date = None) -> str:
    today    = reference_date or date.today()
    pm       = prev_month(today)
    pm_year  = pm.year
    pm_month = pm.month
    pm_nama  = BULAN_NAMA[pm_month]

    fy_year  = fiscal_year_of(pm)
    fy_start = date(fy_year, FY_START_MONTH, 1)

    # ── Bagian 1: Bulan terakhir ──────────────────────────────────────────────
    key  = (pm_year, pm_month)
    data = monthly.get(key, {})
    inc  = data.get('income',  0.0) / 1_000_000   # → Jt
    exp  = data.get('expense', 0.0) / 1_000_000
    sur  = inc - exp

    if sur >= 0:
        verdict1 = f"🟢 *SURPLUS +{fmt(sur)}*"
    else:
        verdict1 = f"🔴 *DEFISIT ({fmt(abs(sur))})*"

    # ── Bagian 2: YTD tabel ───────────────────────────────────────────────────
    ytd_rows = []
    ytd_data = []
    ytd_inc  = 0.0
    ytd_exp  = 0.0
    cur = fy_start
    while (cur.year, cur.month) <= (pm_year, pm_month):
        k  = (cur.year, cur.month)
        d  = monthly.get(k, {})
        mi = d.get('income',  0.0) / 1_000_000
        me = d.get('expense', 0.0) / 1_000_000
        ms = mi - me
        ytd_inc += mi
        ytd_exp += me
        bln_short = BULAN_NAMA[cur.month][:3]
        yr_short  = str(cur.year)[2:]
        icon      = "🟢" if ms >= 0 else "🔴"
        ms_str    = f"+{fmt(ms)}" if ms >= 0 else f"({fmt(abs(ms))})"
        ytd_rows.append(f"{bln_short}'{yr_short}  {fmt(mi)}  {fmt(me)}  {icon}{ms_str}")
        ytd_data.append((bln_short, cur.year, mi, me))
        cur = next_month(cur)

    ytd_sur   = ytd_inc - ytd_exp
    ytd_str   = f"+{fmt(ytd_sur)}" if ytd_sur >= 0 else f"({fmt(abs(ytd_sur))})"
    ytd_icon  = "🟢" if ytd_sur >= 0 else "🔴"
    ytd_label = "SURPLUS" if ytd_sur >= 0 else "DEFISIT"

    # ── Waktu WITA ────────────────────────────────────────────────────────────
    now_wita = datetime.now(timezone(timedelta(hours=8))).strftime('%d %b %Y %H:%M')

    # ── Susun pesan ───────────────────────────────────────────────────────────
    lines = [
        f"📊 *LAPORAN KAS BULANAN*",
        f"*{COMPANY_NAME}*",
        f"━━━━━━━━━━━━━━━━━",
        f"",
        f"*1️⃣ Periode: {pm_nama} {pm_year}*",
        f"💰 Masuk  : *{fmt(inc)}*",
        f"💸 Keluar : *{fmt(exp)}*",
        verdict1,
        f"",
        f"━━━━━━━━━━━━━━━━━",
        f"*2️⃣ YTD Fiskal FY{fy_year}*",
        f"_(Apr {fy_year} – {pm_nama} {pm_year})_",
        f"",
        f"Bulan   Masuk  Keluar  Sur/Def",
        f"──────────────────────",
    ]
    lines += ytd_rows
    lines += [
        f"",
        f"💰 Masuk : *{fmt(ytd_inc)}*",
        f"💸 Keluar: *{fmt(ytd_exp)}*",
        f"{ytd_icon} *{ytd_label} {ytd_str}*",
        f"",
        f"━━━━━━━━━━━━━━━━━",
        f"*3️⃣ Analisa*",
        build_analisa(ytd_data),
        f"",
        f"━━━━━━━━━━━━━━━━━",
        f"🤖 _Auto-report | {now_wita} WITA_",
        f"📁 _Laporan Kas Adhitama_",
    ]
    return '\n'.join(lines)

# ─── Send via Fonnte ──────────────────────────────────────────────────────────
def send_wa(message: str) -> dict:
    log.info("Sending WhatsApp message via Fonnte...")
    resp = requests.post(
        'https://api.fonnte.com/send',
        headers={'Authorization': FONNTE_TOKEN},
        data={
            'target':      WA_TARGET,
            'message':     message,
            'countryCode': '62',
        },
        timeout=30,
    )
    result = {}
    try:
        result = resp.json()
    except Exception:
        result = {'raw': resp.text, 'status_code': resp.status_code}
    log.info(f"Fonnte response: {result}")
    return result

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("  KAS OTOMATIS — start")
    log.info("=" * 60)

    try:
        xlsx_path = download_spreadsheet()
        monthly   = extract_monthly_data(xlsx_path)
        message   = build_wa_message(monthly)
        log.info(f"\n{'─'*60}\n{message}\n{'─'*60}")

        result = send_wa(message)
        if result.get('status') is True or result.get('status') == 'true':
            log.info("✅ WhatsApp message sent successfully.")
        else:
            log.warning(f"⚠️  Fonnte returned: {result}")

        try:
            os.unlink(xlsx_path)
        except Exception:
            pass

        log.info("KAS OTOMATIS — done ✅")
        return True

    except Exception as e:
        log.error(f"FATAL ERROR: {e}\n{traceback.format_exc()}")
        return False


if __name__ == '__main__':
    ok = main()
    sys.exit(0 if ok else 1)
