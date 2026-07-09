#!/usr/bin/env python3
"""
Sales Monitoring Automation - Adhitama
Usage:
  python3 sales_reminder.py morning      # 09:30 Senin-Jumat
  python3 sales_reminder.py evening      # 17:00 Senin-Jumat
  python3 sales_reminder.py weekly       # 11:00 Sabtu
  python3 sales_reminder.py morning --dry-run
"""
import sys
import io
import os
import requests
import openpyxl
from datetime import datetime, date, timedelta
from pathlib import Path

FONNTE_TOKEN = os.environ["FONNTE_TOKEN"]
BASE_DIR = Path(__file__).parent
WEB_TOKEN = "Bearer " + os.environ["ADHITAMA_WEB_TOKEN"]
WEB_API = "https://apps.adhitama.id/api"

SALES_CONFIG = {
    "Surabaya": {
        "nama": "Pipit", "grup_id": "120363423882448709@g.us",
        "rekap_file": "Final Rekap Projek - PIPIT (NA) - SURABAYA.xlsx",
        "monitoring_file": "Sales_Monitoring_Adhitama_v3_Surabaya.xlsx",
        "web_user_id": 23,
        "onedrive_monitoring": os.environ.get("ONEDRIVE_SURABAYA_MONITORING", ""),
        "onedrive_rekap":      os.environ.get("ONEDRIVE_SURABAYA_REKAP", ""),
    },
    "Manado": {
        "nama": "Nayah", "grup_id": "120363402270321041@g.us",
        "rekap_file": "Final Rekap Projek - NUR INAYAH (NI) - MANADO.xlsx",
        "monitoring_file": "Sales_Monitoring_Adhitama_v3_Manado.xlsx",
        "web_user_id": 22,
        "onedrive_monitoring": os.environ.get("ONEDRIVE_MANADO_MONITORING", ""),
        "onedrive_rekap":      os.environ.get("ONEDRIVE_MANADO_REKAP", ""),
    },
    "Makassar": {
        "nama": "Qalbi", "grup_id": "120363406350666051@g.us",
        "rekap_file": "Final Rekap Projek - NURUL QALBI (NQ) - MAKASSAR.xlsx",
        "monitoring_file": "Sales_Monitoring_Adhitama_v3_Makassar.xlsx",
        "web_user_id": 21,
        "onedrive_monitoring": os.environ.get("ONEDRIVE_MAKASSAR_MONITORING", ""),
        "onedrive_rekap":      os.environ.get("ONEDRIVE_MAKASSAR_REKAP", ""),
    },
}

HARI_ID = {
    "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu",
    "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu", "Sunday": "Minggu",
}
BULAN_ID = {
    "January": "Januari", "February": "Februari", "March": "Maret",
    "April": "April", "May": "Mei", "June": "Juni", "July": "Juli",
    "August": "Agustus", "September": "September", "October": "Oktober",
    "November": "November", "December": "Desember",
}
STAGE_LABEL = {
    "A": "On Process / PO", "B": "Deal / TTD", "C": "Negosiasi Klien",
    "D": "Ada Respon Klien", "E": "Quotation",
}
STAGE_EMOJI = {
    "A": "\U0001f3c6", "B": "\U0001f535", "C": "\U0001f7e1",
    "D": "\U0001f7e0", "E": "\U0001f534",
}
SUMBER_DATA = (
    "\n--------------------\n"
    "_Sumber Data:_\n"
    "_1. File Assessment (Rekap Proyek)_\n"
    "_2. File Sales Monitoring_\n"
    "_3. Aktivitas di Web adhitama.id_"
)


def load_wb(local_path, onedrive_url=None):
    """Load workbook: coba OneDrive dulu, fallback ke file lokal."""
    if onedrive_url:
        try:
            r = requests.get(onedrive_url, timeout=20)
            if r.status_code == 200 and r.content[:4] == b'PK\x03\x04':
                return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        except Exception as e:
            print("  [WARN] OneDrive download gagal, pakai file lokal: " + str(e))
    return openpyxl.load_workbook(str(local_path), read_only=True, data_only=True)


def id_date(fmt, dt=None):
    s = (dt or datetime.now()).strftime(fmt)
    for en, id_ in {**HARI_ID, **BULAN_ID}.items():
        s = s.replace(en, id_)
    return s


def get_pipeline(rekap_file, onedrive_url=None):
    result = {
        "per_stage": {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0},
        "win_bulan_ini": 0, "etd_dekat": [], "etd_terlewat": [],
        "total_aktif": 0, "project_names": [],
    }
    today = date.today()
    try:
        wb = load_wb(rekap_file, onedrive_url)
        data_sheet = next((s for s in wb.sheetnames if "GUIDELINE" not in s.upper()), wb.sheetnames[0])
        ws = wb[data_sheet]
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c).strip().upper() if c else "" for c in next(rows_iter)]

        def col(name, fallback=-1):
            for i, h in enumerate(header):
                if name.upper() in h:
                    return i
            return fallback

        i_status = col("STATUS", 1)
        i_stage  = col("STAGE", 2)
        i_prj    = col("PROJECT NAME", 7)
        i_etd    = col("ETD", 11)

        for row in rows_iter:
            if not any(row[:5]):
                continue
            status = str(row[i_status]).strip().upper() if i_status >= 0 and i_status < len(row) and row[i_status] else ""
            stage  = str(row[i_stage]).strip().upper()  if i_stage  >= 0 and i_stage  < len(row) and row[i_stage]  else ""
            prj    = str(row[i_prj]).strip()            if i_prj    >= 0 and i_prj    < len(row) and row[i_prj]    else "-"
            etd    = row[i_etd] if i_etd >= 0 and i_etd < len(row) else None

            if status == "WIN" and etd and isinstance(etd, datetime):
                if etd.month == today.month and etd.year == today.year:
                    result["win_bulan_ini"] += 1

            if status == "ACTIVE" and stage in result["per_stage"]:
                result["per_stage"][stage] += 1
                result["total_aktif"] += 1
                result["project_names"].append(prj.upper())
                if etd and isinstance(etd, datetime):
                    days_left = (etd.date() - today).days
                    if 0 <= days_left <= 30:
                        result["etd_dekat"].append({
                            "nama": prj[:45], "etd": etd.strftime("%d %b"),
                            "days": days_left, "stage": stage,
                        })
                    elif days_left < 0:
                        result["etd_terlewat"].append({
                            "nama": prj[:45], "etd": etd.strftime("%d %b"),
                            "overdue": abs(days_left), "stage": stage,
                        })

        result["etd_dekat"].sort(key=lambda x: x["days"])
        result["etd_terlewat"].sort(key=lambda x: x["overdue"])
        wb.close()
    except Exception as e:
        print("  [WARN] Rekap: " + str(e))
    return result


def in_rekap(mention, project_names):
    """Cek nama proyek di Assessment. Minimal kata 6 huruf untuk hindari false positive."""
    if not mention:
        return False
    words = [w for w in mention.upper().split() if len(w) >= 6]
    if not words:
        return False
    for prj in project_names:
        if any(w in prj for w in words):
            return True
    return False


def cek_assessment(new_quo, fu_quo, customer, project_names):
    """Cek Assessment bertingkat: New Quo -> FU Quo -> Customer name."""
    return (
        in_rekap(new_quo or "", project_names) or
        in_rekap(fu_quo or "", project_names) or
        in_rekap(customer or "", project_names)
    )


def fmt_nilai(nilai):
    if not nilai:
        return ""
    try:
        return " -- Rp " + "{:,.0f}".format(float(nilai)).replace(",", ".")
    except Exception:
        return ""


def get_laporan_harian(monitoring_file, target_date, onedrive_url=None):
    entries = []
    bulan_map = {
        "januari": 1, "februari": 2, "maret": 3, "april": 4,
        "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12,
    }
    try:
        wb = load_wb(monitoring_file, onedrive_url)
        target_month = target_date.month
        sheet_name = None
        for sh in wb.sheetnames:
            if "Laporan" not in sh:
                continue
            sh_lower = sh.lower()
            for nama_bulan, nomor in bulan_map.items():
                if nama_bulan in sh_lower and nomor == target_month:
                    sheet_name = sh
                    break
            if sheet_name:
                break
        if not sheet_name:
            wb.close()
            return entries
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        header_idx = next((i for i, r in enumerate(all_rows) if r[1] and "Tanggal" in str(r[1])), -1)
        if header_idx < 0:
            wb.close()
            return entries
        for row in all_rows[header_idx + 1:]:
            dt = row[1]
            if not isinstance(dt, datetime):
                continue
            if dt.date() == target_date:
                customer = str(row[3]).strip() if row[3] else None
                followup = str(row[6]).strip() if row[6] else None
                quo_baru = str(row[7]).strip() if row[7] else None
                nilai    = row[8]
                if customer or followup or quo_baru:
                    entries.append({
                        "customer": customer,
                        "followup": followup,
                        "quo_baru": quo_baru,
                        "nilai": nilai,
                    })
        wb.close()
    except Exception as e:
        print("  [WARN] Laporan Harian: " + str(e))
    return entries


def get_web_aktivitas(user_id, target_date):
    hasil = []
    try:
        url = WEB_API + "/sales-activities?page=1&limit=200&role_id=1&user_id=" + str(user_id)
        resp = requests.get(url, headers={"Authorization": WEB_TOKEN}, timeout=15)
        if resp.status_code != 200:
            return hasil
        for item in resp.json().get("data", []):
            tgl_str = item.get("tanggal", "")
            try:
                item_date = datetime.fromisoformat(tgl_str.replace("Z", "+00:00")).date()
            except Exception:
                continue
            if item_date == target_date:
                foto = item.get("foto", [])
                hasil.append({"judul": item.get("judul", ""), "ada_foto": len(foto) > 0, "foto_count": len(foto)})
    except Exception as e:
        print("  [WARN] Web API: " + str(e))
    return hasil


def build_pipeline_section(kota, pipeline):
    stage_lines = ""
    for s in ["A", "B", "C", "D", "E"]:
        n = pipeline["per_stage"][s]
        if n:
            stage_lines += STAGE_EMOJI[s] + " Stage " + s + " (" + STAGE_LABEL[s] + "): *" + str(n) + " proyek*\n"
    if not stage_lines:
        stage_lines = "Belum ada proyek aktif\n"

    win_bi   = pipeline["win_bulan_ini"]
    win_line = "✅ WIN bulan ini: *" + str(win_bi) + " proyek*" if win_bi else "\U0001f3af Belum ada WIN bulan ini -- ayo kejar!"

    etd_block = ""
    if pipeline["etd_dekat"]:
        etd_block = "\n⚠️ *ETD <= 30 hari:*\n"
        for p in pipeline["etd_dekat"][:4]:
            etd_block += "  - [" + p["stage"] + "] " + p["nama"] + " -- " + p["etd"] + " (" + str(p["days"]) + " hr lagi)\n"

    overdue_block = ""
    if pipeline["etd_terlewat"]:
        overdue_block = "\n\U0001f6a8 *ETD TERLEWAT - belum update status:*\n"
        for p in pipeline["etd_terlewat"][:4]:
            overdue_block += "  - [" + p["stage"] + "] " + p["nama"] + " -- ETD " + p["etd"] + " (" + str(p["overdue"]) + " hr lalu)\n"

    return (
        "\U0001f4ca *[STATUS PROYEK " + kota.upper() + "]*\n\n"
        + stage_lines + "\n"
        + win_line + etd_block + overdue_block
    ).rstrip()


def build_aktivitas_section(cfg, target_date, label_hari):
    entries   = get_laporan_harian(BASE_DIR / cfg["monitoring_file"], target_date, cfg.get("onedrive_monitoring"))
    web_acts  = get_web_aktivitas(cfg["web_user_id"], target_date)
    pipeline  = get_pipeline(BASE_DIR / cfg["rekap_file"], cfg.get("onedrive_rekap"))
    prj_names = pipeline["project_names"]

    tgl_str = id_date("%A, %d %B %Y", datetime.combine(target_date, datetime.min.time()))
    lines = ["\U0001f4cb *[CEK " + label_hari.upper() + " (" + tgl_str + ")]*", ""]

    # ── Kunjungan ──────────────────────────────────────────────
    kunjungan = [e for e in entries if e.get("customer")]
    if kunjungan:
        lines.append("*\U0001f6b6 Kunjungan:*")
        for e in kunjungan:
            lines.append("  • " + e["customer"][:50])
    else:
        lines.append("*\U0001f6b6 Kunjungan:* _(tidak ada)_")

    lines.append("")

    # ── FU Quo ─────────────────────────────────────────────────
    fu_list = [e for e in entries if e.get("followup")]
    if fu_list:
        lines.append("*\U0001f4cc FU Quo:*")
        for e in fu_list:
            ada = cek_assessment(e["followup"], None, e.get("customer"), prj_names)
            icon = "✅" if ada else "⚠️"
            text = "ada di Assessment" if ada else "belum ada di Assessment!"
            lines.append("  • " + e["followup"][:50] + fmt_nilai(e.get("nilai")) + " → " + icon + " _" + text + "_")
    else:
        lines.append("*\U0001f4cc FU Quo:* _(tidak ada)_")

    lines.append("")

    # ── New Quo ────────────────────────────────────────────────
    boq_list = [e for e in entries if e.get("quo_baru")]
    if boq_list:
        lines.append("*\U0001f4dd New Quo:*")
        for e in boq_list:
            ada = cek_assessment(e["quo_baru"], e.get("followup"), e.get("customer"), prj_names)
            icon = "✅" if ada else "⚠️"
            text = "ada di Assessment" if ada else "belum ada di Assessment!"
            lines.append("  • " + e["quo_baru"][:50] + fmt_nilai(e.get("nilai")) + " → " + icon + " _" + text + "_")
    else:
        lines.append("*\U0001f4dd New Quo:* _(tidak ada)_")

    lines.append("")

    # ── Web Aktivitas ──────────────────────────────────────────
    if web_acts:
        dengan_foto = [a for a in web_acts if a["ada_foto"]]
        tanpa_foto  = [a for a in web_acts if not a["ada_foto"]]
        web_icon = "✅" if dengan_foto else "⚠️"
        web_text = web_icon + " " + str(len(dengan_foto)) + " aktivitas dengan foto"
        if tanpa_foto:
            web_text += ", " + str(len(tanpa_foto)) + " tanpa foto"
    else:
        web_text = "❌ Belum ada upload aktivitas"
    lines.append("*\U0001f4f1 Web Aktivitas:* " + web_text)

    return "\n".join(lines)


def build_weekly_aktivitas(cfg, jumat_lalu):
    """Jumat pekan lalu s/d Kamis pekan ini (6 hari kerja, skip Minggu)."""
    OFFSETS = [0, 1, 3, 4, 5, 6]
    lines = ["\U0001f4c5 *AKTIVITAS PEKAN INI*", ""]
    for off in OFFSETS:
        hari_date = jumat_lalu + timedelta(days=off)
        entries  = get_laporan_harian(BASE_DIR / cfg["monitoring_file"], hari_date, cfg.get("onedrive_monitoring"))
        web_acts = get_web_aktivitas(cfg["web_user_id"], hari_date)

        hari_nama = id_date("%A", datetime.combine(hari_date, datetime.min.time()))
        tgl_str   = hari_date.strftime("%d/%m")
        lines.append("*" + hari_nama + " (" + tgl_str + ")*")

        kunjungan = [e for e in entries if e.get("customer")]
        if kunjungan:
            lines.append("\U0001f6b6 Kunjungan:")
            for e in kunjungan:
                lines.append("  • " + e["customer"][:50])
        else:
            lines.append("\U0001f6b6 Kunjungan: _(tidak ada)_")

        fu_list = [e for e in entries if e.get("followup")]
        if fu_list:
            lines.append("\U0001f4cc FU Quo:")
            for e in fu_list:
                lines.append("  • " + e["followup"][:50] + fmt_nilai(e.get("nilai")))
        else:
            lines.append("\U0001f4cc FU Quo: _(tidak ada)_")

        boq_list = [e for e in entries if e.get("quo_baru")]
        if boq_list:
            lines.append("\U0001f4dd New Quo:")
            for e in boq_list:
                lines.append("  • " + e["quo_baru"][:50] + fmt_nilai(e.get("nilai")))
        else:
            lines.append("\U0001f4dd New Quo: _(tidak ada)_")

        if web_acts:
            dengan_foto = [a for a in web_acts if a["ada_foto"]]
            web_icon = "✅" if len(dengan_foto) == len(web_acts) else "⚠️"
            lines.append("\U0001f4f1 Web: " + web_icon)
            for a in web_acts:
                foto_ket = "✅ ada foto" if a["ada_foto"] else "⚠️ tanpa foto"
                lines.append("  • " + (a["judul"] or "-")[:50] + " -- " + foto_ket)
        else:
            lines.append("\U0001f4f1 Web: ❌ _(tidak ada)_")

        lines.append("")
    return "\n".join(lines).rstrip()


def msg_morning(kota, cfg):
    pipeline  = get_pipeline(BASE_DIR / cfg["rekap_file"], cfg.get("onedrive_rekap"))
    nama      = cfg["nama"]
    tgl       = id_date("%A, %d %B %Y")
    yesterday = date.today() - timedelta(days=1)
    if yesterday.weekday() == 6:
        yesterday -= timedelta(days=2)
    elif yesterday.weekday() == 5:
        yesterday -= timedelta(days=1)

    pipeline_sec  = build_pipeline_section(kota, pipeline)
    aktivitas_sec = build_aktivitas_section(cfg, yesterday, "KEMARIN")

    return (
        "\U0001f305 *Selamat pagi, " + nama + "!*\n"
        "_" + tgl + "_\n\n"
        + pipeline_sec + "\n"
        "\n--------------------\n"
        + aktivitas_sec + "\n\n"
        "_Semangat " + nama + "! \U0001f4aa_"
        + SUMBER_DATA
    ).strip()


def msg_evening(kota, cfg):
    pipeline  = get_pipeline(BASE_DIR / cfg["rekap_file"], cfg.get("onedrive_rekap"))
    nama      = cfg["nama"]
    tgl       = id_date("%A, %d %B %Y")
    today     = date.today()

    pipeline_sec  = build_pipeline_section(kota, pipeline)
    aktivitas_sec = build_aktivitas_section(cfg, today, "HARI INI")

    return (
        "\U0001f306 *Selamat sore, " + nama + "!*\n"
        "_" + tgl + "_\n\n"
        + pipeline_sec + "\n"
        "\n--------------------\n"
        + aktivitas_sec + "\n\n"
        "_Lengkapi data sebelum tutup hari ya! \U0001f64f_"
        + SUMBER_DATA
    ).strip()


def msg_weekly(kota, cfg):
    pipeline = get_pipeline(BASE_DIR / cfg["rekap_file"], cfg.get("onedrive_rekap"))
    bulan    = id_date("%B %Y")

    today = date.today()
    # Berpatok ke Kamis terdekat (akhir window), bukan Jumat -- supaya tetap
    # benar kalau script dijalankan manual di hari selain Jumat. Di hari Jumat
    # asli, ini otomatis sama dengan "today - 7" seperti biasa.
    hari_sejak_kamis = (today.weekday() - 3) % 7  # Kamis = weekday() 3
    kamis_akhir = today - timedelta(days=hari_sejak_kamis)
    jumat_lalu  = kamis_akhir - timedelta(days=6)

    win_bi   = pipeline["win_bulan_ini"]
    win_line = "✅ *" + str(win_bi) + " WIN* bulan ini" if win_bi else "❌ *Belum ada WIN* bulan ini"

    stage_lines = ""
    for s in ["A", "B", "C", "D", "E"]:
        n = pipeline["per_stage"][s]
        if n:
            stage_lines += STAGE_EMOJI[s] + " Stage " + s + " (" + STAGE_LABEL[s] + "): " + str(n) + "\n"
    if not stage_lines:
        stage_lines = "Belum ada proyek aktif\n"

    etd_block = ""
    if pipeline["etd_dekat"]:
        etd_block = "\n⚠️ *ETD Dekat:*\n"
        for p in pipeline["etd_dekat"][:5]:
            etd_block += "  - [" + p["stage"] + "] " + p["nama"] + " -- " + p["etd"] + "\n"

    overdue_block = ""
    if pipeline["etd_terlewat"]:
        overdue_block = "\n\U0001f6a8 *ETD Terlewat - perlu update:*\n"
        for p in pipeline["etd_terlewat"][:5]:
            overdue_block += "  - [" + p["stage"] + "] " + p["nama"] + " (sejak " + p["etd"] + ")\n"

    aktivitas_sec = build_weekly_aktivitas(cfg, jumat_lalu)

    pipeline_sec = (
        "\U0001f4ca *Ringkasan Pipeline*\n\n"
        + win_line + "\n"
        "\U0001f4e6 Total aktif: *" + str(pipeline["total_aktif"]) + " proyek*\n\n"
        "*Pipeline per stage:*\n"
        + stage_lines
        + etd_block
        + overdue_block
    ).rstrip()

    return (
        "\U0001f4ca *REKAP MINGGUAN -- " + kota.upper() + "*\n"
        "_" + bulan + "_\n\n"
        + aktivitas_sec + "\n"
        "\n--------------------\n"
        + pipeline_sec + "\n\n"
        "_Terima kasih! \U0001f64f_"
        + SUMBER_DATA
    ).strip()


def send_wa(target, message):
    try:
        r = requests.post(
            "https://api.fonnte.com/send",
            headers={"Authorization": FONNTE_TOKEN},
            data={"target": os.environ.get("WA_TEST_OVERRIDE") or target, "message": message, "countryCode": "62"},
            timeout=30,
        )
        return r.json()
    except Exception as e:
        return {"status": False, "error": str(e)}


def main():
    args     = sys.argv[1:]
    dry_run  = "--dry-run" in args
    types    = [a for a in args if not a.startswith("--")]
    msg_type = types[0] if types else "morning"

    today = datetime.now()
    if msg_type in ("morning", "evening") and today.weekday() >= 5:
        print("Weekend -- skip.")
        return

    print("[" + today.strftime("%Y-%m-%d %H:%M") + "] Type: " + msg_type + " | Dry-run: " + str(dry_run))
    print("=" * 60)

    for kota, cfg in SALES_CONFIG.items():
        print("\n>> " + kota + " (" + cfg["nama"] + ")")
        if msg_type == "morning":
            message = msg_morning(kota, cfg)
        elif msg_type == "evening":
            message = msg_evening(kota, cfg)
        elif msg_type == "weekly":
            message = msg_weekly(kota, cfg)
        else:
            print("  [ERROR] Unknown type: " + msg_type)
            continue
        print(message)
        print()
        if not dry_run:
            result = send_wa(cfg["grup_id"], message)
            ok = result.get("status", False)
            print("  Fonnte: " + ("OK" if ok else "GAGAL") + " -- " + str(result))
        else:
            print("  [DRY-RUN] Tidak dikirim.")

    print("\nSelesai!")


if __name__ == "__main__":
    main()
